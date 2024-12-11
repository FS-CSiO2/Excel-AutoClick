import openpyxl  # 用于读取Excel文件
import time      # 用于暂停执行
import cv2       # 用于图像处理
import numpy as np  # 用于数组操作
import pyautogui  # 用于模拟鼠标和键盘操作
import os         # 用于文件和路径操作
import pyperclip  # 用于剪贴板操作
import webbrowser # 用于打开网页

def ctrlV(i = 0.1): # 模拟粘贴键盘动作，默认间隔0.1秒
    time.sleep(i) 
    pyautogui.keyDown('ctrl') 
    time.sleep(i) 
    pyautogui.press('v')
    time.sleep(i)   
    pyautogui.keyUp('ctrl')
    time.sleep(i)   

def foundImg(path):# 查找图片返回中心点   
    template = cv2.imread(path) # 读取模板图像  
    _, w, h= template.shape[::-1] # 获取模板图像的w宽度和h高度    
    screen = pyautogui.screenshot() # 获取屏幕截图   
    screen2 = cv2.cvtColor(np.array(screen), cv2.COLOR_RGB2BGR) # 将屏幕截图转换为OpenCV图像
    result = cv2.matchTemplate(screen2, template, cv2.TM_CCOEFF_NORMED) # 在屏幕截图中进行模板匹配
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result) # 获取匹配结果中最大值的位置
    center = (max_loc[0] + (w / 2)  , max_loc[1] + (h / 2)) # 计算匹配位置的中心点
    threshold = 0.92 # 设置匹配阈值
    if max_val > threshold: # 如果匹配值大于阈值，则返回中心点
        return center

def ColorClickOnce(path,LorR): # 点击一次图片（找不到就跳过）
    time.sleep(0.1)  
    center = foundImg(path) # 传递路径给查找图片返回中心点   
    if center is not None: # 如果找到图片，则模拟鼠标左键点击
        if LorR == "L":
            pyautogui.click(center)
        elif LorR == "R":
            pyautogui.click(center,button='right')
        else:
            print("LorR参数错误")
    else:
        print(f"未找到图片:{path}，跳过")

def ColorClickLoop(path,LorR): #点击一次图片（找不到就一直找）
    while True:
        time.sleep(0.1)       
        center = foundImg(path) # 传递路径给查找图片返回中心点       
        if center is not None: # 如果找到图片，则模拟鼠标左键点击
            if LorR == "L":
                pyautogui.click(center)
                break
            elif LorR == "R":
                pyautogui.click(center,button='right')
                break
            else:
                print("LorR参数错误")

def ColorClickMain():
    print("==================开始执行==================")
    row = 2  # 从第二行开始，因为第一行是标题行，逐行读取并执行函数，直到读取到空行为止
    while True:      
        cell_value = worksheet.cell(row=row, column=3).value  # c列是第3列，column的值为3；# 读取c列的值，作为行为类型
        if cell_value is None: # 如果读取到c列空行，则退出循环
            print("==================结束执行==================")
            break
        cell_value1 = worksheet.cell(row=row, column=2).value # b列是第2列，column的值为2；读取b列的值，作为参数1
        if cell_value1 is None: # 如果读取到b列空行，则退出循环 
            print(f"第{row}行b列的内容为空,无参数\n==================结束执行==================")
            break
        if cell_value in [2, 2.1, 3, 3.1, 5, 6, 7, 8, 9, 10] and not isinstance(cell_value1, str): # 当动作需要用到字符串时，如果参数1不是字符串，将其转换为字符串
            cell_value1 = str(cell_value1)
        if cell_value == 1 : #1的时候，等待时间
            print(f"第 {row} 行c列的内容: {cell_value}，行为：等待时间\n b列的内容: {cell_value1}") 
            time.sleep(float(cell_value1)) 
        elif cell_value in [2, 2.1, 3, 3.1]: # 2，2.1，3，3.1的时候，左键（或右键）点击图片一次 （找不到就跳过） 或 （找不到就一直找）
            LorR = 'R' if cell_value in [2.1, 3.1] else 'L'
            ColorClickPicture = cell_value1.replace('\\', "\\\\") #调整好路径里的斜杠           
            if cell_value in [2, 2.1]: # 2（或2.1）的时候，左键（或右键）点击图片一次（找不到就跳过）
                print(f"第 {row} 行c列的内容: {cell_value}，行为：鼠标 {LorR} 键点击一次图片（找不到就跳过）\n b列的内容: {cell_value1}")
                ColorClickOnce(ColorClickPicture, LorR)
            elif cell_value in [3, 3.1]: # 3（或3.1）的时候，左键（或右键）点击图片一次（找不到就一直找）
                print(f"第 {row} 行c列的内容: {cell_value}，行为：鼠标 {LorR} 键点击一次图片（找不到就一直找）\n b列的内容: {cell_value1}")
                ColorClickLoop(ColorClickPicture, LorR)
        elif cell_value == 4 : #4的时候，左键点击坐标
            print(f"第 {row} 行c列的内容: {cell_value}，行为：左键点击坐标\n b列的内容: {cell_value1}")            
            x, y = map(int, cell_value1.split(', ')) # 将字符串转换为坐标        
            pyautogui.click(x, y) # 模拟鼠标左键点击坐标
        elif cell_value == 4.1 : #4.1的时候，右键点击坐标
            print(f"第 {row} 行c列的内容: {cell_value}，行为：右键点击坐标\n b列的内容: {cell_value1}")            
            x, y = map(int, cell_value1.split(', ')) # 将字符串转换为坐标        
            pyautogui.click(x, y,button='right') # 模拟鼠标右键点击坐标
        elif cell_value == 5 : #5的时候，输入文字
            print(f"第 {row} 行c列的内容: {cell_value}，行为：输入文字\n b列的内容: {cell_value1}") 
            pyperclip.copy(cell_value1)
            ctrlV(0.1)
        elif cell_value == 6 : #6的时候，长按一个按键
            print(f"第 {row} 行c列的内容: {cell_value}，行为：长按一个按键\n b列的内容: {cell_value1}") 
            pyautogui.keyDown(cell_value1)
            time.sleep(0.1)
        elif cell_value == 7 : #7的时候，松开一个按键
            print(f"第 {row} 行c列的内容: {cell_value}，行为：松开一个按键\n b列的内容: {cell_value1}")
            pyautogui.keyUp(cell_value1)
        elif cell_value == 8 : #8的时候，快速按一下一个按键
            print(f"第 {row} 行c列的内容: {cell_value}，行为：快速按一下一个按键\n b列的内容: {cell_value1}")
            pyautogui.press(cell_value1)
        elif cell_value == 9 : #9的时候，打开一个网页
            print(f"第 {row} 行c列的内容: {cell_value}，行为：打开一个网页\n b列的内容: {cell_value1}")
            webbrowser.open(cell_value1)
        elif cell_value == 10 : #10的时候，打开一个文件
            print(f"第 {row} 行c列的内容: {cell_value}，行为：打开一个文件\n b列的内容: {cell_value1}")           
            path = cell_value1.replace('\\', "\\\\") #调整好路径里的斜杠
            os.startfile(path)
        elif cell_value in [11, 11.1] : #11(或11.1)的时候，如果b列图片存在，就左键（或右键）点击参数2图片
            LorR = 'R' if cell_value == 11.1 else 'L'
            cell_value2 = worksheet.cell(row=row, column=4).value # d列是第2列，column的值为2；读取b列的值，作为参数2 
            if cell_value2 is None: # 如果读取到d列为空，则退出循环 
                print(f"第{row}行d列的内容为空,无参数\n==================结束执行==================")
                break
            ColorClickPicture = cell_value1.replace('\\', "\\\\") #调整好路径里的斜杠  
            ColorClickPicture1 = cell_value2.replace('\\', "\\\\") #调整好路径里的斜杠
            print(f"第 {row} 行c列的内容: {cell_value}，行为：如果b列图片存在，就鼠标 {LorR} 键点击图片\n b列的内容: {cell_value1} \n d列的内容: {cell_value2}")
            PictuerFound = foundImg(ColorClickPicture) # 传递路径给查找图片返回中心点   
            if PictuerFound is not None: # 如果找到图片，则模拟鼠标点击
                ColorClickLoop(ColorClickPicture1, LorR)
            else :
                print(f"找不到图片{ColorClickPicture1}！")
        else :
            print(f"第 {row} 行c列的内容: {cell_value} 无法识别行为类型") 
        row += 1

current_file_path = os.path.abspath(__file__) # 获取当前脚本文件的绝对路径
current_dir = os.path.dirname(current_file_path) # 获取当前脚本文件所在的目录
os.chdir(current_dir) # 更改当前工作目录为脚本文件所在的目录
print("当前工作目录:", os.getcwd()) # 验证当前工作目录是否更改成功
workbook = openpyxl.load_workbook('.\\Work.xlsx') # 打开表格
worksheet = workbook.active # 选择工作表
ColorClickMain() # 调用主函数