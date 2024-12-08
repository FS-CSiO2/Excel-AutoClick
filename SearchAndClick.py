import openpyxl
import time
import cv2
import numpy as np
import pyautogui
import os

def ColorClickLoop(path):
    while True:
        # 等待0.1秒，避免过于频繁的屏幕截图和模板匹配
        time.sleep(0.1)
        # 读取模板图像
        template = cv2.imread(path)
        # 获取模板图像的w宽度和h高度
        _, w, h= template.shape[::-1]
        # 获取屏幕截图
        screen = pyautogui.screenshot(region=(0, 0, 1920, 1080))
        # 将屏幕截图转换为OpenCV图像
        screen2 = cv2.cvtColor(np.array(screen), cv2.COLOR_RGB2BGR)
        # 在屏幕截图中进行模板匹配
        result = cv2.matchTemplate(screen2, template, cv2.TM_CCOEFF_NORMED)
        # 获取匹配结果中最大值的位置
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        # 计算匹配位置的中心点
        center = (max_loc[0] + (w / 2)  , max_loc[1] + (h / 2))
        # 设置匹配阈值
        threshold = 0.92
        # 如果匹配值大于阈值，模拟鼠标左键点击
        if max_val > threshold:
            pyautogui.click(center)
            break

def ColorClickOnce(path):
    # 读取模板图像
    template = cv2.imread(path)
    # 获取模板图像的w宽度和h高度
    _, w, h= template.shape[::-1]
    # 获取屏幕截图
    screen = pyautogui.screenshot(region=(0, 0, 1920, 1080))
    # 将屏幕截图转换为OpenCV图像
    screen2 = cv2.cvtColor(np.array(screen), cv2.COLOR_RGB2BGR)
    # 在屏幕截图中进行模板匹配
    result = cv2.matchTemplate(screen2, template, cv2.TM_CCOEFF_NORMED)
    # 获取匹配结果中最大值的位置
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    # 计算匹配位置的中心点
    center = (max_loc[0] + (w / 2)  , max_loc[1] + (h / 2))
    # 设置匹配阈值
    threshold = 0.92
    # 如果匹配值大于阈值，模拟鼠标左键点击
    if max_val > threshold:
        pyautogui.click(center)

# 获取当前脚本文件的绝对路径
current_file_path = os.path.abspath(__file__)

# 获取当前脚本文件所在的目录
current_dir = os.path.dirname(current_file_path)

# 更改当前工作目录为脚本文件所在的目录
os.chdir(current_dir)

# 验证当前工作目录是否更改成功
print("当前工作目录:", os.getcwd())

# 打开表格
workbook = openpyxl.load_workbook('.\\Work.xlsx')

# 选择工作表
worksheet = workbook.active

# 从第二行开始逐行读取并执行函数，直到读取到空行为止
row = 2  # 从第二行开始，因为第一行是标题行
while True:
    cell_value = worksheet.cell(row=row, column=3).value  # c列是第3列，即column的值   
    if cell_value is None: # 如果读取到空行，则退出循环
        break           
    cell_value2 = worksheet.cell(row=row, column=2).value # b列是第2列，即column的值
    if cell_value == 1 : #1的时候，点击图片（找不到就跳过）
        print(f"第 {row} 行c列的内容: {cell_value}，动作：点击一次图片（找不到就跳过）") 
        ColorClickPicture = cell_value2.replace('\\', "\\\\")
        print(f"第 {row} 行b列的点击图片路径: {ColorClickPicture}")
        ColorClickOnce(ColorClickPicture)
    elif cell_value == 2 : #2的时候，点击图片（找不到就一直找）
        print(f"第 {row} 行c列的内容: {cell_value}，动作：点击一次图片（找不到就一直找）") 
        ColorClickPicture = cell_value2.replace('\\', "\\\\")
        print(f"第 {row} 行b列的点击图片路径: {ColorClickPicture}")
        ColorClickLoop(ColorClickPicture)
    elif cell_value == 3 : #3的时候，等待时间
        print(f"第 {row} 行c列的内容: {cell_value}，动作：等待时间") 
        print(f"第 {row} 行b列的等待时间: {cell_value2}")
        time.sleep(cell_value2)
    else :
        print(f"第 {row} 行c列的内容: {cell_value}无法识别") 
    row += 1
